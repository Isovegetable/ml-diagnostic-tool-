"""
生成发布包中的所有文件（PDF / HTML / XLSX / TXT）
"""
import os, shutil, re, sys
from datetime import datetime

OUT = "发布包"

# ============================================================
# 工具：PDF 基础
# ============================================================
from fpdf import FPDF

_CJK_FONT = None
for _p in [
    "C:/Windows/Fonts/simhei.ttf", "C:/Windows/Fonts/deng.ttf", "C:/Windows/Fonts/msyh.ttc",
    "/System/Library/Fonts/PingFang.ttc", "/System/Library/Fonts/STHeiti Light.ttc",
    "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
    "/usr/share/fonts/truetype/wqy/wqy-microhei.ttc",
]:
    if os.path.exists(_p):
        _CJK_FONT = _p
        break


class PDF(FPDF):
    def __init__(self, title="", watermark=""):
        super().__init__()
        self._watermark = watermark
        self._use_cjk = _CJK_FONT is not None and self._add_cjk(_CJK_FONT)
        if title:
            self.set_title(title)

    def _add_cjk(self, path):
        try:
            is_ttc = path.lower().endswith(".ttc")
            k = "CJK"
            if is_ttc:
                self.add_font(k, "", path, ttc_index=0)
                self.add_font(k, "B", path, ttc_index=0)
            else:
                self.add_font(k, "", path)
                self.add_font(k, "B", path)
            return True
        except Exception:
            return False

    def _f(self, style="", size=10):
        return ("CJK", style if style in ("", "B") else "", size) if self._use_cjk else ("Helvetica", style, size)

    def footer(self):
        self.set_y(-15)
        if self._watermark:
            self.set_font(*self._f("", 7))
            self.set_text_color(180, 180, 180)
            self.cell(0, 10, self._watermark, align="C")
        elif self.page_no() > 1:
            self.set_font(*self._f("", 8))
            self.set_text_color(150, 150, 150)
            self.cell(0, 10, "- {} -".format(self.page_no()), align="C")


# ============================================================
# 1. 01_产品介绍.pdf
# ============================================================
def make_product_intro():
    pdf = PDF("产品介绍", "懂点AI的C学长")
    pdf.add_page()

    pdf.ln(25)
    pdf.set_font(*pdf._f("B", 26))
    pdf.set_text_color(31, 119, 180)
    pdf.cell(0, 14, "材料机器学习自动诊断工具", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)
    pdf.set_font(*pdf._f("", 12))
    pdf.set_text_color(80, 80, 80)
    pdf.cell(0, 8, "Materials ML Diagnostic Tool", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(8)

    pdf.set_draw_color(31, 119, 180)
    pdf.line(30, pdf.get_y(), 180, pdf.get_y())
    pdf.ln(8)

    def heading(text):
        pdf.set_font(*pdf._f("B", 14))
        pdf.set_text_color(31, 119, 180)
        pdf.cell(0, 10, text, new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)

    def body(text):
        pdf.set_font(*pdf._f("", 10))
        pdf.set_text_color(50, 50, 50)
        pdf.multi_cell(0, 6.5, text)
        pdf.ln(3)

    heading("产品简介")
    body("本工具专为材料科学、化学、能源领域的科研人员设计。"
         "上传实验数据的 Excel 或 CSV 文件，即可自动完成机器学习建模、模型评估、特征分析和诊断报告输出。"
         "无需编程基础，小白也能轻松上手。")

    heading("核心功能")
    for title, desc in [
        ("一键建模", "6 种机器学习模型自动训练，自动选出最佳模型"),
        ("智能诊断", "A/B/C/D 四级评分，告诉你数据适不适合做预测"),
        ("可解释分析", "SHAP 图、PDP 图，知道每个特征怎么影响结果"),
        ("报告导出", "HTML / PDF 格式报告，可打印、可分享"),
        ("完全本地", "数据不上传服务器，断网也能用，保障数据安全"),
    ]:
        pdf.set_font(*pdf._f("B", 10))
        pdf.set_text_color(50, 50, 50)
        pdf.cell(35, 7, "  " + title)
        pdf.set_font(*pdf._f("", 10))
        pdf.set_text_color(100, 100, 100)
        pdf.cell(0, 7, desc, new_x="LMARGIN", new_y="NEXT")
        pdf.ln(1)

    pdf.ln(3)
    heading("适用场景")
    body("  * 材料性能预测(强度、导电率、硬度等)\n"
         "  * 实验数据快速建模与可行性判断\n"
         "  * 特征重要性分析，指导下一步实验方向\n"
         "  * 学术论文数据分析和图表制作")

    heading("技术栈")
    body("Python / Streamlit / scikit-learn / Optuna / SHAP / XGBoost / LightGBM / CatBoost")

    heading("隐私安全")
    body("本工具完全在本地运行，所有数据不离开你的电脑。\n"
         "适合未发表的论文数据和企业配方数据。建议不使用工具后及时删除上传的数据文件。")

    pdf.ln(5)
    pdf.set_draw_color(200, 200, 200)
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(3)
    pdf.set_font(*pdf._f("", 8))
    pdf.set_text_color(140, 140, 140)
    pdf.cell(0, 5, "生成日期：" + datetime.now().strftime('%Y-%m-%d'), align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 5, "版权所有 (c) 懂点AI的C学长", align="C")
    pdf.output(os.path.join(OUT, "01_产品介绍.pdf"))
    print("[OK] 01_产品介绍.pdf")


# ============================================================
# 2. 02_样例报告.html + 02_样例报告.pdf
# ============================================================
def make_sample_report():
    # --- HTML ---
    html = """<!DOCTYPE html>
<html lang="zh-CN">
<head><meta charset="utf-8"><title>样例报告 - 混凝土强度预测</title>
<style>
body { font-family: 'Microsoft YaHei', sans-serif; max-width: 900px; margin: 40px auto; padding: 20px; color: #333; }
h1 { color: #1F77B4; border-bottom: 3px solid #1F77B4; padding-bottom: 10px; }
h2 { color: #1F77B4; margin-top: 30px; }
.level-A { background: #e8f5e9; border-left: 5px solid #1a7d36; padding: 10px 15px; font-size: 18px; font-weight: bold; }
.metric { display: inline-block; background: #eef2f7; border-radius: 6px; padding: 10px 18px; margin: 5px; text-align: center; }
.metric .val { font-size: 22px; font-weight: bold; color: #1F77B4; }
.metric .lab { font-size: 12px; color: #888; }
table { border-collapse: collapse; width: 100%; margin: 10px 0; }
th, td { border: 1px solid #ddd; padding: 8px 12px; text-align: left; }
th { background: #1F77B4; color: white; }
tr:nth-child(even) { background: #f9f9f9; }
</style></head><body>
<h1>材料机器学习诊断报告</h1>
<p><strong>目标列：</strong>compress_strength | <strong>生成时间：</strong>2026-06-26</p>
<p><strong>数据集：</strong>混凝土强度数据（15 样本, 4 特征）</p>
<h2>诊断等级</h2>
<div class="level-A">A  适合建模  评分 82/100</div>
<p>数据质量良好，特征与目标变量相关性较强，适合进行机器学习建模。</p>
<h2>最佳模型：Random Forest</h2>
<div>
<div class="metric"><div class="val">0.9234</div><div class="lab">CV R2</div></div>
<div class="metric"><div class="val">0.8978</div><div class="lab">Test R2</div></div>
<div class="metric"><div class="val">0.9512</div><div class="lab">Train R2</div></div>
</div>
<h2>模型对比</h2>
<table>
<tr><th>模型</th><th>CV R2</th><th>Test R2</th><th>排名</th></tr>
<tr><td>Random Forest</td><td>0.9234</td><td>0.8978</td><td>1</td></tr>
<tr><td>Gradient Boosting</td><td>0.9012</td><td>0.8756</td><td>2</td></tr>
<tr><td>XGBoost</td><td>0.8895</td><td>0.8612</td><td>3</td></tr>
<tr><td>LightGBM</td><td>0.8721</td><td>0.8435</td><td>4</td></tr>
<tr><td>线性回归</td><td>0.8345</td><td>0.8123</td><td>5</td></tr>
<tr><td>CatBoost</td><td>0.8245</td><td>0.7988</td><td>6</td></tr>
</table>
<h2>建议</h2>
<ul>
<li>样本量偏少（15 条），建议补充更多实验数据</li>
<li>交叉验证 R2 标准差较低，模型稳定性较好</li>
<li>训练集和测试集 R2 差距在合理范围内，无明显过拟合</li>
</ul>
<p style="color:#999; font-size:12px; margin-top:40px;">本报告由材料机器学习自动诊断工具生成，仅供参考</p>
</body></html>"""

    with open(os.path.join(OUT, "02_样例报告.html"), "w", encoding="utf-8") as f:
        f.write(html)
    print("[OK] 02_样例报告.html")

    # --- PDF ---
    pdf = PDF("样例报告", "懂点AI的C学长")
    pdf.add_page()
    pdf.ln(15)
    pdf.set_font(*pdf._f("B", 22))
    pdf.set_text_color(31, 119, 180)
    pdf.cell(0, 12, "材料机器学习诊断报告", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(5)
    pdf.set_font(*pdf._f("", 10))
    pdf.set_text_color(100, 100, 100)
    pdf.cell(0, 7, "目标列：compress_strength    数据集：混凝土强度数据(15样本,4特征)", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(10)
    pdf.set_fill_color(26, 125, 54)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font(*pdf._f("B", 24))
    pdf.cell(25, 25, "  A  ", fill=True, align="C")
    pdf.set_font(*pdf._f("B", 12))
    pdf.set_text_color(26, 125, 54)
    pdf.cell(0, 25, "  适合建模    评分 82/100", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(5)
    pdf.set_font(*pdf._f("", 10))
    pdf.set_text_color(80, 80, 80)
    pdf.multi_cell(0, 6, "数据质量良好，特征与目标变量相关性较强，适合进行机器学习建模。")
    pdf.ln(8)
    pdf.set_font(*pdf._f("B", 14))
    pdf.set_text_color(31, 119, 180)
    pdf.cell(0, 9, "最佳模型：Random Forest", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)
    x0, y0 = pdf.get_x(), pdf.get_y()
    for i, (lab, val) in enumerate([("CV R2", "0.9234"), ("Test R2", "0.8978"), ("Train R2", "0.9512")]):
        x = x0 + i * 62
        pdf.set_xy(x, y0)
        pdf.set_fill_color(240, 244, 248)
        pdf.rect(x, y0, 58, 22, style="DF")
        pdf.set_xy(x + 5, y0 + 3)
        pdf.set_font(*pdf._f("B", 16))
        pdf.set_text_color(31, 119, 180)
        pdf.cell(48, 8, val, align="C")
        pdf.set_xy(x + 5, y0 + 12)
        pdf.set_font(*pdf._f("", 8))
        pdf.set_text_color(100, 100, 100)
        pdf.cell(48, 6, lab, align="C")
    pdf.set_xy(x0, y0 + 28)
    pdf.set_font(*pdf._f("B", 11))
    pdf.set_text_color(50, 50, 50)
    pdf.cell(0, 8, "模型对比", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)
    pdf.set_fill_color(31, 119, 180)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font(*pdf._f("B", 9))
    for name, w in [("模型", 60), ("CV R2", 30), ("Test R2", 30), ("排名", 20)]:
        pdf.cell(w, 7, " " + name, fill=True)
    pdf.ln()
    pdf.set_font(*pdf._f("", 9))
    for i, (name, cv, test, rank) in enumerate([
        ("Random Forest", 0.9234, 0.8978, 1),
        ("Gradient Boosting", 0.9012, 0.8756, 2),
        ("XGBoost", 0.8895, 0.8612, 3),
        ("LightGBM", 0.8721, 0.8435, 4),
        ("线性回归", 0.8345, 0.8123, 5),
        ("CatBoost", 0.8245, 0.7988, 6),
    ]):
        pdf.set_fill_color(248, 248, 248) if i % 2 == 0 else pdf.set_fill_color(255, 255, 255)
        pdf.set_text_color(50, 50, 50)
        pdf.cell(60, 6.5, " " + name, fill=True)
        pdf.cell(30, 6.5, "{:.4f}".format(cv), fill=True, align="C")
        pdf.cell(30, 6.5, "{:.4f}".format(test), fill=True, align="C")
        pdf.cell(20, 6.5, str(rank), fill=True, align="C")
        pdf.ln()
    pdf.ln(5)
    pdf.set_font(*pdf._f("", 9))
    pdf.set_text_color(100, 100, 100)
    pdf.multi_cell(0, 5, "建议：样本量偏少(15条)，建议补充更多实验数据。")
    pdf.output(os.path.join(OUT, "02_样例报告.pdf"))
    print("[OK] 02_样例报告.pdf")


# ============================================================
# 3. 03_使用说明.pdf
# ============================================================
def make_user_guide():
    src = "使用说明.md"
    with open(src, "r", encoding="utf-8") as f:
        md = f.read()

    pdf = PDF("使用说明", "懂点AI的C学长")
    pdf.set_auto_page_break(auto=True, margin=20)
    pdf.add_page()

    lines = md.split("\n")
    i = 0
    in_code, code_text = False, ""

    while i < len(lines):
        line = lines[i]
        if line.startswith("```"):
            if not in_code:
                in_code, code_text = True, ""
                i += 1
                continue
            else:
                in_code = False
                if code_text.strip():
                    pdf.set_font(*pdf._f("", 9))
                    pdf.set_text_color(80, 80, 80)
                    pdf.set_fill_color(245, 245, 245)
                    pdf.multi_cell(0, 5.5, code_text.strip(), fill=True)
                    pdf.ln(3)
                i += 1
                continue
        if in_code:
            code_text += line + "\n"
            i += 1
            continue

        s = line.strip()
        if not s:
            pdf.ln(2)
            i += 1
            continue
        if s == "---":
            pdf.set_draw_color(200, 200, 200)
            pdf.line(20, pdf.get_y(), 190, pdf.get_y())
            pdf.ln(4)
            i += 1
            continue
        if s.startswith("# ") and not s.startswith("## "):
            pdf.set_font(*pdf._f("B", 18))
            pdf.set_text_color(31, 119, 180)
            pdf.multi_cell(0, 9, s[2:].strip())
            pdf.ln(3)
            i += 1
            continue
        if s.startswith("## ") and not s.startswith("### "):
            text = s[3:].strip()
            clean = text.split(" ", 1)[-1] if text and ord(text[0]) > 0xFFFF else text
            pdf.set_font(*pdf._f("B", 14))
            pdf.set_text_color(31, 119, 180)
            pdf.cell(0, 9, clean, new_x="LMARGIN", new_y="NEXT")
            pdf.ln(2)
            i += 1
            continue
        if s.startswith("### "):
            pdf.set_font(*pdf._f("B", 11))
            pdf.set_text_color(50, 100, 150)
            pdf.cell(0, 7.5, s[4:].strip(), new_x="LMARGIN", new_y="NEXT")
            pdf.ln(1)
            i += 1
            continue
        if s.startswith("> "):
            t = s[2:].strip()
            pdf.set_font(*pdf._f("I", 10))
            pdf.set_text_color(100, 80, 50)
            pdf.multi_cell(0, 6.5, t)
            pdf.ln(1)
            i += 1
            continue
        if s.startswith("- ") or s.startswith("* "):
            t = s.lstrip(" -").strip()
            t = re.sub(r"\*\*(.*?)\*\*", r"\1", t)
            pdf.set_font(*pdf._f("", 10))
            pdf.set_text_color(50, 50, 50)
            pdf.cell(6, 6.5, "")
            pdf.multi_cell(0, 6.5, t)
            pdf.ln(0.5)
            i += 1
            continue
        # 普通段落
        t = re.sub(r"\*\*(.*?)\*\*", r"\1", s)
        pdf.set_font(*pdf._f("", 10))
        pdf.set_text_color(50, 50, 50)
        pdf.multi_cell(0, 6.5, t)
        pdf.ln(1)
        i += 1

    pdf.output(os.path.join(OUT, "03_使用说明.pdf"))
    print("[OK] 03_使用说明.pdf")


# ============================================================
# 4. 04_示例数据.xlsx
# ============================================================
def make_sample_data():
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "混凝土强度数据"
    headers = ["cement", "water", "superplastic", "age", "compress_strength"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill(start_color="1F77B4", end_color="1F77B4", fill_type="solid")
        cell.alignment = openpyxl.styles.Alignment(horizontal="center")
    for r, row in enumerate([
        (540, 162, 2.5, 28, 79.99), (380, 165, 0.0, 90, 61.89),
        (490, 170, 3.0, 14, 55.20), (420, 175, 1.5, 28, 48.50),
        (510, 160, 2.0, 56, 72.30), (450, 168, 2.2, 7, 38.75),
        (600, 155, 3.5, 28, 85.40), (350, 180, 0.0, 90, 45.60),
        (470, 172, 1.8, 14, 52.10), (530, 158, 2.8, 56, 78.90),
        (400, 178, 1.2, 28, 52.30), (560, 150, 4.0, 180, 96.50),
        (390, 182, 0.5, 7, 32.10), (480, 166, 2.0, 28, 65.40),
        (440, 170, 1.5, 56, 60.20),
    ], 2):
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.alignment = openpyxl.styles.Alignment(horizontal="center")
    wb.save(os.path.join(OUT, "04_示例数据.xlsx"))
    print("[OK] 04_示例数据.xlsx")


# ============================================================
# 5. 06_常见问题.pdf
# ============================================================
def make_faq():
    pdf = PDF("常见问题", "懂点AI的C学长")
    pdf.add_page()
    pdf.set_font(*pdf._f("B", 20))
    pdf.set_text_color(31, 119, 180)
    pdf.cell(0, 12, "常见问题解答", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(5)

    faqs = [
        ("双击 start.bat 后黑色窗口一闪就没了？",
         "可能是 Python 没装好。\n在文件夹地址栏输入 cmd 回车，在黑框里输入 start.bat，能看到具体错误。"),
        ("提示【Python 未检测到】？",
         "安装 Python 时没勾「Add Python to PATH」。\n重新运行安装程序，记得勾上再装一次。"),
        ("浏览器打开后是白屏或显示「无法访问此网站」？",
         "刷新一下页面(F5)，或手动输入 http://localhost:8501"),
        ("训练时卡住不动了？",
         "超过 5 分钟没反应：关掉窗口，重新双击 start.bat"),
        ("结果 R2 很低，诊断等级是 C 或 D？",
         "不代表工具坏了。可能样本太少、数据噪音大、目标变量没有明显规律。"),
        ("我的数据会不会被上传？",
         "完全本地运行，断网也能用。数据不会离开你的电脑。"),
        ("支持哪些文件格式？",
         ".xlsx / .xls / .csv。行=样本，列=特征。"),
        ("Mac 用户怎么使用？",
         "阅读同目录下的「03_使用说明.pdf」，有完整 Mac 步骤。"),
    ]

    for q, a in faqs:
        if pdf.get_y() > 240:
            pdf.add_page()
        pdf.set_font(*pdf._f("B", 11))
        pdf.set_text_color(31, 119, 180)
        pdf.multi_cell(0, 7, "Q: " + q)
        pdf.ln(1)
        pdf.set_font(*pdf._f("", 10))
        pdf.set_text_color(80, 80, 80)
        for para in a.split("\n"):
            pdf.multi_cell(0, 6, "   " + para)
            pdf.ln(0.5)
        pdf.ln(4)

    pdf.output(os.path.join(OUT, "06_常见问题.pdf"))
    print("[OK] 06_常见问题.pdf")


# ============================================================
# 6. 07_联系作者.txt
# ============================================================
def make_contact():
    text = """感谢您关注「材料机器学习自动诊断工具」！

━━━━━━━━━━━━━━━━━━━━━━━━
  联系作者
━━━━━━━━━━━━━━━━━━━━━━━━

本工具目前处于测试阶段，欢迎免费体验。

  作者：懂点AI的C学长
  微信：mt196297057

  功能建议、问题反馈、内测交流均可联系。

━━━━━━━━━━━━━━━━━━━━━━━━

后续更新会在网盘同步，欢迎关注！
"""

    with open(os.path.join(OUT, "07_联系作者.txt"), "w", encoding="utf-8") as f:
        f.write(text)
    print("[OK] 07_联系作者.txt")


# ============================================================
# 执行
# ============================================================
if __name__ == "__main__":
    os.makedirs(OUT, exist_ok=True)
    # 清理旧文件
    for f in os.listdir(OUT):
        fp = os.path.join(OUT, f)
        if os.path.isfile(fp):
            os.remove(fp)

    make_product_intro()
    make_sample_report()
    make_user_guide()
    make_sample_data()
    make_faq()
    make_contact()

    print("\n=== 全部生成完毕！文件在「{}」目录中 ===".format(OUT))
    for f in sorted(os.listdir(OUT)):
        fp = os.path.join(OUT, f)
        if os.path.isfile(fp):
            s = os.path.getsize(fp)
            print("  {}  ({:.1f} KB)".format(f, s/1024) if s < 1024*1024 else "  {}  ({:.0f} MB)".format(f, s/1024/1024))
